# -*- coding: utf-8 -*-
"""
Joins Graubünden vote results («Keine 10-Millionen-Schweiz!», 14.06.2026) with
provisional 2025 foreign-resident-share data per municipality, and writes a
compact data.json for the scatter-plot chart in index.html.

Sources (hardcoded paths, see README context in graubuenden-stats/CLAUDE.md
conventions: prep scripts in this project always hardcode source paths):
  - Vote results:      20260614_Statistikdaten_Abstimmungen.xlsx (Kanton Graubünden)
  - Population/Aus-
    laenderanteil:     News_Staendige Wohnbevoelkerung ... 2025p (2).xlsx (BFS STATPOP)
"""
import glob
import json
import math
import os
import re
import sys

import openpyxl

DOWNLOADS = r"C:\Users\julir\Downloads"
VOTE_PATH = os.path.join(DOWNLOADS, "20260614_Statistikdaten_Abstimmungen.xlsx")
POP_PATH_PRIMARY = os.path.join(
    DOWNLOADS, "News_Ständige Wohnbevölkerung nach Staatsangehörigkeit und Geschlecht 2025p (2).xlsx"
)
OUT_DIR = os.path.dirname(os.path.abspath(__file__))


def resolve_pop_path():
    """Use the exact known filename; fall back to a glob match if the
    download was renamed (e.g. missing the "(2)" duplicate-download suffix)."""
    if os.path.exists(POP_PATH_PRIMARY):
        return POP_PATH_PRIMARY
    candidates = glob.glob(os.path.join(DOWNLOADS, "News_Ständige Wohnbevölkerung*2025p*.xlsx"))
    if len(candidates) == 1:
        print(f"[WARN] Primary population file not found, using fallback match: {candidates[0]}")
        return candidates[0]
    raise FileNotFoundError(
        f"Population file not found at {POP_PATH_PRIMARY!r} and fallback glob found "
        f"{len(candidates)} candidates (expected exactly 1): {candidates}"
    )


def read_vote_data(path):
    """Returns dict: gemeinde -> {region, ja_anteil, stimmbeteiligung}.
    Vorlage 1 («Keine 10-Millionen-Schweiz!») only. Layout verified directly:
    sheet 1, data rows 9-108 (100 Gemeinden), row 109 = cantonal "Total" (excluded).
    Col 1=Gemeinde, 2=Region, 11=Ja-Stimmen in %, 13=Stimmbet. in %.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    vorlage_raw = ws.cell(row=4, column=1).value or ""
    vorlage_name = re.sub(r"^Vorlage \d+:\s*\([^)]*\)\s*", "", vorlage_raw).strip()

    data = {}
    for r in range(9, 109):  # rows 9..108 inclusive
        name = ws.cell(row=r, column=1).value
        if name is None:
            continue
        if name == "Total":
            raise AssertionError(f"Unexpected 'Total' row inside data range at row {r}")
        region = ws.cell(row=r, column=2).value
        ja_anteil = ws.cell(row=r, column=11).value
        stimmbeteiligung = ws.cell(row=r, column=13).value
        data[name] = {
            "region": region,
            "ja_anteil": float(ja_anteil),
            "stimmbeteiligung": float(stimmbeteiligung),
        }

    # confirm row 109 really is the cantonal total we expect to exclude
    total_row_name = ws.cell(row=109, column=1).value
    if total_row_name != "Total":
        raise AssertionError(f"Expected row 109 to be 'Total', got {total_row_name!r}")

    if len(data) != 100:
        raise AssertionError(f"Expected 100 Gemeinden in vote file, got {len(data)}")

    return data, vorlage_name


def read_population_data(path):
    """Returns dict: gemeinde -> {bevoelkerung, auslaenderanteil}, plus the
    cantonal total population for a reconciliation check.
    Layout verified directly: sheet "2025p", data rows 16-127 inclusive.
    Row 16 = "GRAUBÜNDEN" cantonal total (excluded), 11 "Region ..." aggregate
    rows interspersed (excluded), remaining 100 rows = Gemeinden.
    Col 1=Gemeinde, 2=Total_Total, 8=Ausländer/innen_Total (by position, not
    exact umlaut string, since terminal/console display of this column header
    is unreliable - the underlying file bytes are correct UTF-8).
    Rows 129-140 repeat the canton+region summary block and must NOT be read.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["2025p"]

    canton_total = None
    data = {}
    for r in range(16, 128):  # rows 16..127 inclusive
        name = ws.cell(row=r, column=1).value
        if name is None:
            continue
        total = ws.cell(row=r, column=2).value
        auslaender = ws.cell(row=r, column=8).value
        if name == "GRAUBÜNDEN":
            canton_total = total
            continue
        if name.startswith("Region "):
            continue
        anteil = auslaender / total * 100 if total else None
        data[name] = {
            "bevoelkerung": int(total),
            "auslaenderanteil": round(anteil, 2) if anteil is not None else None,
        }

    if canton_total is None:
        raise AssertionError("Could not find 'GRAUBÜNDEN' cantonal total row")
    if len(data) != 100:
        raise AssertionError(f"Expected 100 Gemeinden in population file, got {len(data)}")

    reconciliation = sum(v["bevoelkerung"] for v in data.values())
    if reconciliation != canton_total:
        raise AssertionError(
            f"Reconciliation failed: sum of Gemeinde populations ({reconciliation}) "
            f"!= cantonal total ({canton_total})"
        )

    return data, canton_total


def pearson_and_regression(xs, ys):
    n = len(xs)
    mean_x = sum(xs) / n
    mean_y = sum(ys) / n
    cov = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys)) / n
    var_x = sum((x - mean_x) ** 2 for x in xs) / n
    var_y = sum((y - mean_y) ** 2 for y in ys) / n
    r = cov / math.sqrt(var_x * var_y)
    slope = cov / var_x
    intercept = mean_y - slope * mean_x
    return r, slope, intercept, mean_x, mean_y


def main():
    pop_path = resolve_pop_path()

    vote_data, vorlage_name = read_vote_data(VOTE_PATH)
    pop_data, canton_total = read_population_data(pop_path)

    vote_names = set(vote_data)
    pop_names = set(pop_data)
    if vote_names != pop_names:
        only_vote = sorted(vote_names - pop_names)
        only_pop = sorted(pop_names - vote_names)
        print("[ERROR] Gemeinde name mismatch between sources:")
        print("  only in vote file:", only_vote)
        print("  only in population file:", only_pop)
        sys.exit(1)

    records = []
    for name in sorted(vote_names):
        v = vote_data[name]
        p = pop_data[name]
        for key, val in (("ja_anteil", v["ja_anteil"]), ("auslaenderanteil", p["auslaenderanteil"])):
            if not (0 <= val <= 100):
                raise AssertionError(f"{name}: {key}={val} out of [0,100] range")
        records.append({
            "gemeinde": name,
            "region": v["region"],
            "ja_anteil": v["ja_anteil"],
            "stimmbeteiligung": v["stimmbeteiligung"],
            "auslaenderanteil": p["auslaenderanteil"],
            "bevoelkerung": p["bevoelkerung"],
        })

    xs = [rec["auslaenderanteil"] for rec in records]
    ys = [rec["ja_anteil"] for rec in records]
    r, slope, intercept, mean_x, mean_y = pearson_and_regression(xs, ys)

    out = {
        "meta": {
            "n": len(records),
            "vorlage": vorlage_name,
            "abstimmungsdatum": "2026-06-14",
            "pearson_r": round(r, 3),
            "regression": {"slope": round(slope, 4), "intercept": round(intercept, 3)},
            "mean_auslaenderanteil": round(mean_x, 2),
            "mean_ja_anteil": round(mean_y, 2),
            "einordnung_kurz": "Kein klares Muster erkennbar",
            "einordnung_lang": "Der Ausländeranteil einer Gemeinde sagt für sich allein praktisch nichts über ihr Abstimmungsverhalten aus.",
            "quelle_abstimmung": "Kanton Graubünden, Statistikdaten Abstimmungen, 14.06.2026 (20260614_Statistikdaten_Abstimmungen.xlsx)",
            "quelle_bevoelkerung": "BFS STATPOP via data.gr.ch, Ständige Wohnbevölkerung nach Staatsangehörigkeit, provisorisch, Stand 31.12.2025",
        },
        "gemeinden": records,
    }

    out_json_path = os.path.join(OUT_DIR, "data.json")
    with open(out_json_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    out_js_path = os.path.join(OUT_DIR, "data.js")
    with open(out_js_path, "w", encoding="utf-8") as f:
        f.write("const DATA = ")
        json.dump(out, f, ensure_ascii=False, indent=2)
        f.write(";\n")

    # --- console summary for plausibility check ---
    by_auslaender = sorted(records, key=lambda rec: rec["auslaenderanteil"])
    print(f"OK: {len(records)}/{len(records)} Gemeinden gejoint, Reconciliation OK (Kantonstotal {canton_total}).")
    print(f"Vorlage: {vorlage_name}")
    print(f"Pearson r = {r:.3f}  (slope={slope:.4f}, intercept={intercept:.3f})")
    print(f"Mittelwert Auslaenderanteil = {mean_x:.2f}%   Mittelwert Ja-Anteil = {mean_y:.2f}%")
    print("\nTop 5 Auslaenderanteil:")
    for rec in by_auslaender[-5:]:
        print(f"  {rec['gemeinde']!r}: Auslaenderanteil={rec['auslaenderanteil']}%  Ja={rec['ja_anteil']}%")
    print("\nBottom 5 Auslaenderanteil:")
    for rec in by_auslaender[:5]:
        print(f"  {rec['gemeinde']!r}: Auslaenderanteil={rec['auslaenderanteil']}%  Ja={rec['ja_anteil']}%")
    print(f"\nWritten: {out_json_path}")
    print(f"Written: {out_js_path}")


if __name__ == "__main__":
    main()
