# -*- coding: utf-8 -*-
"""
Joins existing vote data (data.json) with Leerwohnungsziffer (vacant-housing
per capita) from the official Statistik Graubünden source, for the
"Leerwohnungsziffer vs. Ja-Anteil" chart.

WICHTIG: Dieser Zusammenhang ist NICHT robust (Pearson r=0.22 p=0.025 sieht
knapp signifikant aus, bricht aber bei Spearman/Trimming/Kontrolle für
Gemeindegrösse fast vollständig zusammen) - bestätigt sowohl mit der
Kollegen-Tabelle als auch mit dieser offiziellen Quelle (praktisch identische
Werte, Unterschied nur in der 4. Nachkommastelle).

Source: 1006_Leerwohnungsbestände 2010-2025.xlsx (Statistik Graubünden,
BFS Leerwohnungszählung, Stand 1. Juni, letztmals aktualisiert 09.09.2025).
"""
import json
import math
import os

import openpyxl

OUT_DIR = os.path.dirname(os.path.abspath(__file__))
VOTE_JSON = os.path.join(OUT_DIR, "data.json")
LEERWOHN_PATH = r"C:\Users\julir\Downloads\1006_Leerwohnungsbestände 2010-2025.xlsx"

# Namensabweichungen zwischen dieser Quelle und unseren 100 Gemeinden
ALIASES = {
    "St.Moritz": "St. Moritz",
    "Schmitten": "Schmitten (GR)",
    "Sils im Engadin / Segl": "Sils im Engadin/Segl",
    "Roveredo": "Roveredo (GR)",
}


def pearson_and_regression(xs, ys):
    n = len(xs)
    mean_x = sum(xs) / n
    mean_y = sum(ys) / n
    cov = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys)) / n
    var_x = sum((x - mean_x) ** 2 for x in xs) / n
    var_y = sum((y - mean_y) ** 2 for y in ys) / n
    r = cov / math.sqrt(var_x * var_y)
    slope = cov / var_x
    intercept = mean_y - slope * mean_x
    return r, slope, intercept


def read_leerwohnungen():
    """2025-Spalte (Total, Spalte D) je Gemeinde. Layout verified: Zeilen
    15-139, Zeile 15 = 'GRAUBÜNDEN'-Total (ausschliessen), 'Region '-Zeilen
    ausschliessen, Footnote-Zeilen liegen erst ab 141 (Range stoppt vorher)."""
    wb = openpyxl.load_workbook(LEERWOHN_PATH, data_only=True)
    ws = wb["Leerwohnungsbestand"]
    out = {}
    canton_total = None
    for r in range(15, 140):
        name = ws.cell(row=r, column=1).value
        if name is None:
            continue
        name = name.strip()
        if name == "GRAUBÜNDEN":
            canton_total = ws.cell(row=r, column=4).value
            continue
        if name.startswith("Region "):
            continue
        name = ALIASES.get(name, name)
        out[name] = ws.cell(row=r, column=4).value

    if len(out) != 100:
        raise AssertionError(f"Erwartet 100 Gemeinden, erhalten {len(out)}")
    reconciliation = sum(out.values())
    if reconciliation != canton_total:
        raise AssertionError(f"Reconciliation fehlgeschlagen: {reconciliation} != {canton_total}")
    return out


def main():
    with open(VOTE_JSON, encoding="utf-8") as f:
        vote = json.load(f)["gemeinden"]

    leerwohn_count = read_leerwohnungen()

    vote_names = {g["gemeinde"] for g in vote}
    if vote_names != set(leerwohn_count):
        raise AssertionError(f"Namens-Mismatch: {vote_names ^ set(leerwohn_count)}")

    records = []
    for g in vote:
        lw_pro_kopf = leerwohn_count[g["gemeinde"]] / g["bevoelkerung"]
        records.append({
            "gemeinde": g["gemeinde"],
            "region": g["region"],
            "ja_anteil": g["ja_anteil"],
            "leerwohn_kopf": round(lw_pro_kopf, 4),
            "bevoelkerung": g["bevoelkerung"],
        })

    print(f"Gejoint: {len(records)}/{len(vote)} Gemeinden, Reconciliation OK")

    xs = [r["leerwohn_kopf"] for r in records]
    ys = [r["ja_anteil"] for r in records]
    r, slope, intercept = pearson_and_regression(xs, ys)

    out = {
        "meta": {
            "n": len(records),
            "pearson_r": round(r, 3),
            "regression": {"slope": round(slope, 4), "intercept": round(intercept, 3)},
            "einordnung_kurz": "Schwacher Zusammenhang",
            "einordnung_lang": "Pearson r=0.22 (schwach) – bricht bei Rangkorrelation (Spearman) und Kontrolle für "
                                "Gemeindegrösse grösstenteils zusammen.",
            "quelle_leerwohnungen": "Statistik Graubünden / BFS Leerwohnungszählung, Stand 1.6.2025 (1006_Leerwohnungsbestände 2010-2025.xlsx)",
            "quelle_abstimmung": "Kanton Graubünden, Statistikdaten Abstimmungen, 14.06.2026",
        },
        "gemeinden": records,
    }

    out_json = os.path.join(OUT_DIR, "data_leerwohnungen.json")
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    out_js = os.path.join(OUT_DIR, "data_leerwohnungen.js")
    with open(out_js, "w", encoding="utf-8") as f:
        f.write("const DATA_LW = ")
        json.dump(out, f, ensure_ascii=False, indent=2)
        f.write(";\n")

    print(f"Pearson r = {r:.3f}  (slope={slope:.4f}, intercept={intercept:.3f})")
    print(f"Written: {out_json}")
    print(f"Written: {out_js}")


if __name__ == "__main__":
    main()
