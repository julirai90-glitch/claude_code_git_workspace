"""Build gender-data.json: combines 2024 age pyramid (data.gr.ch) + 2025p totals (Excel)
+ historical Geschlechter-Zeitreihen (BFS PxWeb) seit 1981 fuer Kanton, seit 2010 pro Gemeinde.

Sources:
  - C:\\Users\\julir\\Downloads\\dvs_awt_soci_202502111.csv
    = Staendige Wohnbevoelkerung 2024 per Gemeinde x Geschlecht x Altersklasse (data.gr.ch)
  - C:\\Users\\julir\\Downloads\\News_Staendige Wohnbevoelkerung ... 2025p (1).xlsx
    = Provisorische 2025 Zahlen per Gemeinde + CH / Auslaender Split (Statistik GR)
  - BFS PxWeb px-x-0102020000_201 "Demografische Bilanz nach institutionellen Gliederungen"
    = Bestand am 31.12. nach Jahr (1981-2024) x Gemeinde x Geschlecht (Komponente=16)
    Wird fuer Kanton-Zeitreihe ab 1981 und Gemeinde-Zeitreihe ab 2010 benutzt.

Output: ../gender-data.json, consumed by gender-* embeds.

Note zu Fusionen: Gemeindestand 2025 = 100 Gemeinden. Tschiertschen-Praden wurde in
Churwalden fusioniert. Fuer 2024er Pyramide und 2010-2024er Zeitreihe werden die Werte
dieser beiden addiert (FUSIONS_2025). Fuer Gemeinden, die zwischen 2010 und 2024 fusioniert
wurden, beginnt die Zeitreihe ab dem Jahr, in dem die heutige Gemeindenummer existiert.
"""
import csv
import json
import ssl
import urllib.request
from collections import defaultdict
from pathlib import Path

import openpyxl

# ---- Paths ----
BASE = Path(__file__).resolve().parent
OUT_JSON = BASE / "gender-data.json"

CSV_2024 = Path(r"C:\Users\julir\Downloads\dvs_awt_soci_202502111.csv")
XLSX_2025 = Path(r"C:\Users\julir\Downloads\News_Ständige Wohnbevölkerung nach Staatsangehörigkeit und Geschlecht 2025p (1).xlsx")

# ---- Altersklassen (Reihenfolge wichtig: wird als Array im JSON gespeichert) ----
KLASSEN = [
    "0-4 Jahre", "5-9 Jahre", "10-14 Jahre", "15-19 Jahre", "20-24 Jahre",
    "25-29 Jahre", "30-34 Jahre", "35-39 Jahre", "40-44 Jahre", "45-49 Jahre",
    "50-54 Jahre", "55-59 Jahre", "60-64 Jahre", "65-69 Jahre", "70-74 Jahre",
    "75-79 Jahre", "80-84 Jahre", "85-89 Jahre", "90-94 Jahre", "95-99 Jahre",
    "100 Jahre und mehr",
]
MIDPOINTS = {k: (i * 5 + 2) if k != "100 Jahre und mehr" else 102 for i, k in enumerate(KLASSEN)}

# Gemeinden, die in 2025 in Churwalden aufgegangen sind:
FUSIONS_2025 = {"Tschiertschen-Praden": "Churwalden"}

# Mapping historische BFS-Gemeindenummer -> heutige BFS-Gemeindenummer (Stand 2025)
# fuer alle Gemeindenummer-Wechsel zwischen 2010 und 2024 in GR.
# Tschiertschen-Praden 3932 -> Churwalden 3911 (per 2025).
FUSIONS_2025_BY_NR = {3932: 3911}

# ---- BFS PxWeb config ----
PXWEB_URL = "https://www.pxweb.bfs.admin.ch/api/v1/de/px-x-0102020000_201/px-x-0102020000_201.px"
PXWEB_GR_KANTON_CODE = "1201"  # "- Graubuenden / Grigioni / Grischun" in der Geographic-Dimension
PXWEB_KOMPONENTE_BESTAND_31_12 = "16"


def read_2024_pyramids():
    """Liefert dict: gemeinde -> {'nr': gnr, 'klassen_m': {kl: n}, 'klassen_f': {kl: n}}."""
    g = defaultdict(lambda: {"nr": None, "klassen_m": defaultdict(int), "klassen_f": defaultdict(int)})
    with open(CSV_2024, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            if row["Bevölkerungstyp"] != "Ständige Wohnbevölkerung":
                continue
            if row["Jahr"] != "2024":
                continue
            name = row["Gemeinde"]
            klasse = row["Altersklasse"]
            if klasse not in MIDPOINTS:
                continue
            n = int(row["Anzahl Personen"])
            rec = g[name]
            rec["nr"] = int(row["Gemeindenummer"])
            if row["Geschlecht"] == "Mann":
                rec["klassen_m"][klasse] += n
            else:
                rec["klassen_f"][klasse] += n
    return g


def apply_fusions(pyramids):
    """Addiere fusionierte Gemeinden in ihre Nachfolger-Gemeinde."""
    for old, new in FUSIONS_2025.items():
        if old not in pyramids:
            continue
        src = pyramids[old]
        dst = pyramids[new]
        for kl, n in src["klassen_m"].items():
            dst["klassen_m"][kl] += n
        for kl, n in src["klassen_f"].items():
            dst["klassen_f"][kl] += n
        del pyramids[old]


def read_2025p_excel():
    """Liefert:
        - kanton_row: dict fuer GRAUBUENDEN
        - gemeinden_2025: list of dicts {name, total, m, f, ch_total, ch_m, ch_f, au_total, au_m, au_f}
    """
    wb = openpyxl.load_workbook(XLSX_2025, data_only=True)
    ws = wb["2025p"]

    def num(v):
        # Statistik GR schreibt "." fuer "keine Daten"
        if v is None or v == "." or v == "":
            return 0
        try:
            return int(v)
        except (TypeError, ValueError):
            return 0

    kanton = None
    gemeinden = []
    for r in range(16, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        total = ws.cell(row=r, column=2).value
        if not name or total is None:
            continue
        # Fix Umlaut encoding
        name = name.replace("\u00dc", "Ü").strip()
        if name.startswith("(") or name.startswith("Quelle") or name.startswith("Letztmals"):
            continue
        rec = {
            "name": name,
            "total": num(total),
            "m": num(ws.cell(row=r, column=3).value),
            "f": num(ws.cell(row=r, column=4).value),
            "ch_total": num(ws.cell(row=r, column=5).value),
            "ch_m": num(ws.cell(row=r, column=6).value),
            "ch_f": num(ws.cell(row=r, column=7).value),
            "au_total": num(ws.cell(row=r, column=8).value),
            "au_m": num(ws.cell(row=r, column=9).value),
            "au_f": num(ws.cell(row=r, column=10).value),
        }
        if name.startswith("GRAUB"):
            kanton = rec
        elif name.startswith("Region "):
            continue  # Regionen nicht gebraucht
        else:
            gemeinden.append(rec)
    return kanton, gemeinden


def _pxweb_fetch_metadata():
    """Holt die Tabellen-Metadaten und liefert die 3 GR-Code-Listen:
        kanton_code, regionen=[(code, name)], gemeinden=[(code, bfs_nr, name)].
    Verwendet nur stdlib (urllib + ssl context to handle Windows cert issues).
    """
    import re
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    with urllib.request.urlopen(PXWEB_URL, context=ctx) as r:
        meta = json.loads(r.read())
    geo = next(v for v in meta["variables"] if "Kanton" in v["code"])
    codes, texts = geo["values"], geo["valueTexts"]
    gr_start = codes.index(PXWEB_GR_KANTON_CODE)
    gr_end = None
    for i, t in enumerate(texts[gr_start + 1:], start=gr_start + 1):
        if t.startswith("- "):
            gr_end = i
            break
    regions, gemeinden = [], []
    for c, t in list(zip(codes, texts))[gr_start:gr_end]:
        if t.startswith(">> Region"):
            regions.append((c, t.replace(">> ", "").strip()))
        elif t.startswith("......"):
            m = re.match(r"\.+\s*(\d{4})\s+(.+)", t)
            if m:
                gemeinden.append((c, int(m.group(1)), m.group(2).strip()))
    return PXWEB_GR_KANTON_CODE, regions, gemeinden


def _pxweb_query(geo_codes, year_codes):
    """POST-Query an PxWeb. Liefert dict (year, geo_code, sex_code) -> count."""
    body = {
        "query": [
            {"code": "Jahr", "selection": {"filter": "item", "values": year_codes}},
            {"code": "Kanton (-) / Bezirk (>>) / Gemeinde (......)",
             "selection": {"filter": "item", "values": geo_codes}},
            {"code": "Staatsangehörigkeit (Kategorie)",
             "selection": {"filter": "item", "values": ["0"]}},
            {"code": "Geschlecht", "selection": {"filter": "item", "values": ["1", "2"]}},
            {"code": "Demografische Komponente",
             "selection": {"filter": "item", "values": [PXWEB_KOMPONENTE_BESTAND_31_12]}},
        ],
        "response": {"format": "json-stat2"},
    }
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(
        PXWEB_URL,
        data=json.dumps(body).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, context=ctx) as r:
        d = json.loads(r.read())
    # json-stat2: dimensions in 'dimension', values in 'value' (flat array).
    dims = d["id"]            # ordered list of dimension codes
    sizes = d["size"]
    cats = {dim: d["dimension"][dim]["category"]["index"] for dim in dims}
    rev = {dim: {v: k for k, v in cats[dim].items()} for dim in dims}
    values = d["value"]
    # Compute strides for index decoding
    strides = []
    s = 1
    for sz in reversed(sizes):
        strides.insert(0, s)
        s *= sz
    out = {}
    for flat_idx, v in enumerate(values):
        if v is None:
            continue
        coord = []
        idx = flat_idx
        for stride in strides:
            coord.append(idx // stride)
            idx %= stride
        keys = {dims[i]: rev[dims[i]][coord[i]] for i in range(len(dims))}
        year = keys["Jahr"]
        geo_code = keys["Kanton (-) / Bezirk (>>) / Gemeinde (......)"]
        sex = keys["Geschlecht"]
        out[(year, geo_code, sex)] = int(v)
    return out


def fetch_timeseries():
    """Liefert dict mit drei Sektionen:
       {
         "kanton":   [{year, m, f}, ...]  (1981-2024),
         "regionen": {region_name: [{year, m, f}, ...]} (2010-2024),
         "gemeinden":{bfs_nr (str): [{year, m, f}, ...]} (2010-2024, mit Fusionen aggregiert),
       }
    """
    print("Fetching BFS PxWeb metadata ...")
    kanton_code, regions, gemeinden = _pxweb_fetch_metadata()
    print(f"  -> Kanton {kanton_code}, {len(regions)} Regionen, {len(gemeinden)} Gemeinden (historisch)")

    # Gemeinde-BFS-Nr -> Region-Name (anhand Reihenfolge in BFS-Hierarchie)
    gemeinde_to_region = _build_gemeinde_region_map(regions, gemeinden)

    # Query 1: Kanton 1981-2024
    years_full = [str(y) for y in range(1981, 2025)]
    print(f"  Query Kanton 1981-2024 ({len(years_full)} Jahre)...")
    raw_kanton = _pxweb_query([kanton_code], years_full)

    # Query 2: alle Regionen + Gemeinden 2010-2024
    years_short = [str(y) for y in range(2010, 2025)]
    geo_codes = [c for c, _ in regions] + [c for c, _, _ in gemeinden]
    print(f"  Query Regionen+Gemeinden 2010-2024 ({len(geo_codes)} Geo-Einheiten)...")
    raw_geo = _pxweb_query(geo_codes, years_short)

    # Build Kanton series
    kanton_ts = []
    for y in years_full:
        m = raw_kanton.get((y, kanton_code, "1"), 0)
        f = raw_kanton.get((y, kanton_code, "2"), 0)
        if m + f > 0:
            kanton_ts.append({"year": int(y), "m": m, "f": f})

    # Build Regionen series
    regionen_ts = {}
    for code, name in regions:
        series = []
        for y in years_short:
            m = raw_geo.get((y, code, "1"), 0)
            f = raw_geo.get((y, code, "2"), 0)
            if m + f > 0:
                series.append({"year": int(y), "m": m, "f": f})
        regionen_ts[name] = series

    # Build Gemeinden series mit Fusion-Aggregation
    # Schritt A: per (jahr, current_nr) summieren
    nr_to_current = {}  # historische BFS-Nr -> aktuelle (heutige) BFS-Nr
    for hist_nr, curr_nr in FUSIONS_2025_BY_NR.items():
        nr_to_current[hist_nr] = curr_nr

    agg = defaultdict(lambda: defaultdict(int))  # (year, current_nr) -> {'m':..., 'f':...}
    for code, bfs_nr, _ in gemeinden:
        current_nr = nr_to_current.get(bfs_nr, bfs_nr)
        for y in years_short:
            m = raw_geo.get((y, code, "1"), 0)
            f = raw_geo.get((y, code, "2"), 0)
            if m + f > 0:
                agg[(int(y), current_nr)]["m"] += m
                agg[(int(y), current_nr)]["f"] += f

    gemeinden_ts = defaultdict(list)
    # Liste aller heutigen Gemeindenummern (alle, ausser jenen, die fusioniert wurden)
    current_nrs = sorted({nr_to_current.get(bfs_nr, bfs_nr) for _, bfs_nr, _ in gemeinden})
    for nr in current_nrs:
        for y in range(2010, 2025):
            r = agg.get((y, nr))
            if r:
                gemeinden_ts[str(nr)].append({"year": y, "m": r["m"], "f": r["f"]})

    print(f"  -> Kanton: {len(kanton_ts)} Punkte, "
          f"Regionen: {len(regionen_ts)} mit je {len(next(iter(regionen_ts.values())))} Punkten, "
          f"Gemeinden: {len(gemeinden_ts)}")
    return {
        "kanton": kanton_ts,
        "regionen": regionen_ts,
        "gemeinden": dict(gemeinden_ts),
        "_gemeinde_to_region": gemeinde_to_region,
    }


def _build_gemeinde_region_map(regions, gemeinden):
    """Anhand der Reihenfolge der BFS-Geo-Codes (1201, 1202 Region, 1203 Gemeinde, ...)
    wird jede Gemeinde der zuletzt davor stehenden Region zugeordnet.
    regions: [(code, name)], gemeinden: [(code, bfs_nr, name)].
    Gibt {bfs_nr -> region_name} zurueck.
    """
    region_codes = {c: n for c, n in regions}
    # Sortiere nach geo_code (Geo-Codes sind aufsteigend)
    all_entries = [(c, "R", n) for c, n in regions] + [(c, "G", nr) for c, nr, _ in gemeinden]
    all_entries.sort(key=lambda x: int(x[0]))
    out = {}
    current = None
    for code, kind, payload in all_entries:
        if kind == "R":
            current = payload
        elif kind == "G" and current:
            out[payload] = current
    return out


def build_pyramid_pairs(klassen_m, klassen_f):
    """Liefert [[m, f], ...] in KLASSEN-Reihenfolge."""
    return [[klassen_m.get(k, 0), klassen_f.get(k, 0)] for k in KLASSEN]


def compute_avg_age_from_pyramid(pairs):
    total = 0
    summe = 0.0
    for i, (m, f) in enumerate(pairs):
        mid = MIDPOINTS[KLASSEN[i]]
        total += m + f
        summe += (m + f) * mid
    return round(summe / total, 2) if total else None


def main():
    print(f"Reading 2024 age data: {CSV_2024.name}")
    pyr = read_2024_pyramids()
    print(f"  -> {len(pyr)} Gemeinden (before fusion)")
    apply_fusions(pyr)
    print(f"  -> {len(pyr)} Gemeinden (after fusion)")

    print(f"Reading 2025p Excel: {XLSX_2025.name}")
    kanton25, gem25 = read_2025p_excel()
    print(f"  -> Kanton: {kanton25['total']}, {len(gem25)} Gemeinden")

    # Sanity checks: Gemeindeliste 2024 (nach Fusion) == 2025
    names_24 = set(pyr.keys())
    names_25 = {g["name"] for g in gem25}
    missing_24 = names_25 - names_24
    missing_25 = names_24 - names_25
    if missing_24 or missing_25:
        print(f"  WARN only in 2025: {sorted(missing_24)}")
        print(f"  WARN only in 2024: {sorted(missing_25)}")

    # Build Gemeinde-Liste: 2025p totals + 2024 pyramide
    g_list = []
    for g in sorted(gem25, key=lambda x: x["total"], reverse=True):
        name = g["name"]
        py = pyr.get(name)
        if py is None:
            print(f"  SKIP no 2024 pyramide for {name}")
            continue
        pairs = build_pyramid_pairs(py["klassen_m"], py["klassen_f"])
        total25 = g["total"]
        male_pct25 = round(g["m"] / total25 * 100, 2)
        # CH und AU Maennerquoten
        ch_male_pct = round(g["ch_m"] / g["ch_total"] * 100, 2) if g["ch_total"] else None
        au_male_pct = round(g["au_m"] / g["au_total"] * 100, 2) if g["au_total"] else None
        avg_age_2024 = compute_avg_age_from_pyramid(pairs)
        g_list.append({
            "name": name,
            "nr": py["nr"],
            "total_2025p": total25,
            "m_2025p": g["m"],
            "f_2025p": g["f"],
            "male_pct_2025p": male_pct25,
            "ch_total": g["ch_total"],
            "ch_m": g["ch_m"],
            "ch_f": g["ch_f"],
            "ch_male_pct": ch_male_pct,
            "au_total": g["au_total"],
            "au_m": g["au_m"],
            "au_f": g["au_f"],
            "au_male_pct": au_male_pct,
            "avg_age_2024": avg_age_2024,
            "pyramid_2024": pairs,
        })

    # Kanton-Pyramide 2024 (Summe ueber alle Gemeinden)
    kanton_m = defaultdict(int)
    kanton_f = defaultdict(int)
    for rec in pyr.values():
        for kl, n in rec["klassen_m"].items():
            kanton_m[kl] += n
        for kl, n in rec["klassen_f"].items():
            kanton_f[kl] += n
    kanton_pairs = build_pyramid_pairs(kanton_m, kanton_f)
    kanton_avg_age_2024 = compute_avg_age_from_pyramid(kanton_pairs)

    meta = {
        "gr_total_2025p": kanton25["total"],
        "gr_m_2025p": kanton25["m"],
        "gr_f_2025p": kanton25["f"],
        "gr_male_pct_2025p": round(kanton25["m"] / kanton25["total"] * 100, 2),
        "gr_ch_total": kanton25["ch_total"],
        "gr_ch_m": kanton25["ch_m"],
        "gr_ch_f": kanton25["ch_f"],
        "gr_ch_male_pct": round(kanton25["ch_m"] / kanton25["ch_total"] * 100, 2),
        "gr_au_total": kanton25["au_total"],
        "gr_au_m": kanton25["au_m"],
        "gr_au_f": kanton25["au_f"],
        "gr_au_male_pct": round(kanton25["au_m"] / kanton25["au_total"] * 100, 2),
        "gr_avg_age_2024": kanton_avg_age_2024,
        "n_gemeinden": len(g_list),
        "source_2024": "data.gr.ch · dvs_awt_soci_202502111 · Ständige Wohnbevölkerung 2024",
        "source_2025p": "Statistik Graubünden · STATPOP 2025p (provisorisch, Stand 02.04.2025)",
        "gemeindestand": "2025 (100 Gemeinden, Tschiertschen-Praden in Churwalden fusioniert)",
    }

    # ---- Zeitreihen 1981-2024 (BFS PxWeb) + 2025p (Excel) ----
    ts = fetch_timeseries()

    # Append 2025p Punkt an alle Serien
    ts["kanton"].append({"year": 2025, "m": kanton25["m"], "f": kanton25["f"], "p": True})
    name_to_nr = {g["name"]: g["nr"] for g in g_list}
    name_to_2025 = {g["name"]: g for g in g_list}
    for name, g in name_to_2025.items():
        nr = str(name_to_nr[name])
        if nr in ts["gemeinden"]:
            ts["gemeinden"][nr].append({"year": 2025, "m": g["m_2025p"], "f": g["f_2025p"], "p": True})
    # Regionen-2025p = Summe der jeweiligen Gemeinden 2025p
    g2region = ts.pop("_gemeinde_to_region", {})
    region_25 = defaultdict(lambda: {"m": 0, "f": 0})
    for g in g_list:
        # Heutiger Gemeindenummer-Mapping fuer alte (fusionierte) BFS-Nrs
        rname = g2region.get(g["nr"])
        if not rname:
            # Tschiertschen-Praden 3932 -> Churwalden 3911 - g2region hat 3932 als key, nicht 3911
            for hist, curr in FUSIONS_2025_BY_NR.items():
                if curr == g["nr"]:
                    rname = g2region.get(hist); break
        if rname:
            region_25[rname]["m"] += g["m_2025p"]
            region_25[rname]["f"] += g["f_2025p"]
    for rname, vals in region_25.items():
        if rname in ts["regionen"]:
            ts["regionen"][rname].append({"year": 2025, "m": vals["m"], "f": vals["f"], "p": True})

    out = {
        "meta": meta,
        "klassen": KLASSEN,
        "kanton_pyramid_2024": kanton_pairs,
        "gemeinden": g_list,
        "timeseries": ts,
    }

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))

    size_kb = OUT_JSON.stat().st_size / 1024
    print(f"\nWritten {OUT_JSON} ({size_kb:.1f} KB)")
    print(f"\n-- Verifikation --")
    print(f"Kanton 2025p Total: {meta['gr_total_2025p']:,} (expected 207'334)")
    print(f"Kanton 2025p Maennerquote: {meta['gr_male_pct_2025p']}% (expected 50.44)")
    print(f"Kanton 2025p CH Maennerquote: {meta['gr_ch_male_pct']}% (expected 49.31)")
    print(f"Kanton 2025p AU Maennerquote: {meta['gr_au_male_pct']}% (expected 54.64)")
    print(f"Kanton 2024 avg age: {meta['gr_avg_age_2024']} (expected ca. 45.2)")

    # Top 5 maennlichste / weiblichste 2025p
    print(f"\nTop 5 maennlichste (2025p):")
    for g in sorted(g_list, key=lambda x: x["male_pct_2025p"], reverse=True)[:5]:
        print(f"  {g['name']}: {g['male_pct_2025p']}% ({g['total_2025p']} EW)")
    print(f"Top 5 weiblichste (2025p):")
    for g in sorted(g_list, key=lambda x: x["male_pct_2025p"])[:5]:
        print(f"  {g['name']}: {g['male_pct_2025p']}% ({g['total_2025p']} EW)")

    # Zeitreihe-Verifikation
    print(f"\n-- Zeitreihen --")
    k = ts["kanton"]
    if k:
        first = k[0]
        last_full = next((p for p in reversed(k) if not p.get("p")), None)
        last = k[-1]
        def pct(p): return round(p["m"] / (p["m"] + p["f"]) * 100, 2)
        print(f"Kanton {first['year']}: M={first['m']:,} F={first['f']:,} -> {pct(first)}% Maenner")
        if last_full:
            print(f"Kanton {last_full['year']}: M={last_full['m']:,} F={last_full['f']:,} -> {pct(last_full)}% Maenner")
        print(f"Kanton {last['year']}{'p' if last.get('p') else ''}: M={last['m']:,} F={last['f']:,} -> {pct(last)}% Maenner")
    print(f"Anzahl Gemeinden mit Zeitreihe: {len(ts['gemeinden'])}")
    print(f"Anzahl Regionen mit Zeitreihe: {len(ts['regionen'])}")


if __name__ == "__main__":
    main()
